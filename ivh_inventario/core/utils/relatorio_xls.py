from datetime import date

from openpyxl import Workbook

from ivh_inventario.core.utils.email import enviar_email_xls
from ivh_inventario.settings import STATIC_ROOT


def gerar_planilha(model, tipo, usuario, dt_ini=None, dt_fim=None):
    campos = [field.name for field in model.model._meta.get_fields()
              if field.name != 'uuid'
              if field.name != 'saida'
              if field.name != 'entrada'
              if field.name != 'entrada_pai'
              if field.name != 'saida_pai'
              if field.name != 'doc_fisc'
              if field.name != 'is_ultimo']
    wb = Workbook()
    ws = wb.active

    for i, campo in enumerate(campos):
        coluna = i + 1
        ws.cell(row=1, column=coluna, value=campo)

    for i, obj in enumerate(model):
        linha = i + 2

        for j, campo in enumerate(campos):
            if campo == "uuid":
                continue
            coluna = j + 1
            valor = getattr(obj, campo)
            if hasattr(valor, "cnpj_cpf"):
                valor = valor.cnpj_cpf
            if hasattr(valor, "descricao"):
                valor = valor.descricao
            valor = str(valor)
            ws.cell(row=linha, column=coluna, value=valor)

    wb.save(f'{STATIC_ROOT}/arquivo.xlsx')

    if dt_ini is None:
        dt_ini = date.today()
    if dt_fim is None:
        dt_fim = date.today()

    enviar_email_xls(
        arquivo=f'{STATIC_ROOT}/arquivo.xlsx',
        usuario=usuario,
        titulo="[IVH Inventario] teste email",
        html="xls",
        dados={
            "usuario": usuario,
            "tipo": tipo,
            "dt_ini": dt_ini,
            "dt_fim": dt_fim
        })





